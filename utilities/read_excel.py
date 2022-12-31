from openpyxl import Workbook, load_workbook


class ReadExcel:
    book = load_workbook('../testdata/login_info.xlsx')
    sheet = book.active

    def read_excel_phone(self):
        ExcelPhoneNumber = self.sheet.cell(2, 1).value
        return ExcelPhoneNumber

    def read_excel_password(self):
        ExcelPassword = self.sheet.cell(2, 2).value
        return ExcelPassword
